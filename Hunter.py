from pyhunter import PyHunter
from openpyxl import Workbook

# Nombres: Jordan Garcia Bernal y Aaron Adrian Gonzales Alfaro

# Sustituye la siguiente API key por la tuya
# Se cambia el ultimo caracter
api_hunter = PyHunter('21326844b124923a81c807400920286157203d29')


# Funcion para Busqueda
def Busqueda(organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    busqueda = 1
    resultado = api_hunter.domain_search(company=organizacion, limit=busqueda,
                                         emails_type='personal')
    return resultado


# Funcion para guardar informacion
def GuardarInformacion(datosEncontrados, organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    # Se crea el archivo Excel con un nombre
    libro.save("Hunter" + organizacion + ".xlsx")
    # Se llena el archivo Excel con la informacion
    hoja.cell(1, 1, "Dominio")
    hoja.cell(1, 2, datosEncontrados['domain'])
    hoja.cell(2, 1, "Organización")
    hoja.cell(2, 2, datosEncontrados['organization'])
    hoja.cell(3, 1, "Correo")
    hoja.cell(3, 2, datosEncontrados['emails'][0]['value'])
    hoja.cell(4, 1, "Nombre(s)")
    hoja.cell(4, 2, datosEncontrados['emails'][0]['first_name'])
    hoja.cell(5, 1, "Apellidos")
    hoja.cell(5, 2, datosEncontrados['emails'][0]['last_name'])
    libro.save("Hunter" + organizacion + ".xlsx")

def main(organizacion_investigar):
    print("Script para buscar información")
    # organizacion_investigar es la variable que
    # guarda el string de la organizacion a investigar

    #organizacion_investigar = input("Organización a investigar: ")
    datosEncontrados = Busqueda(organizacion_investigar)
    # Verifica que se encuentren datos
    if datosEncontrados is None:
        exit()
    else:
        print(datosEncontrados)
        print(type(datosEncontrados))
        # Se manda informacion para el excel
        GuardarInformacion(datosEncontrados, organizacion_investigar)
